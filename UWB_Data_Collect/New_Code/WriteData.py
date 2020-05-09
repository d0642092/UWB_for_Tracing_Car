#資料輸出
from openpyxl import Workbook, load_workbook

# 拿取自己寫的檔案
from UWB_Data_Collect.New_Code.CalActualDis import *
from UWB_Data_Collect.New_Code.CatchData import *

class writerData(threading.Thread):
    def __init__(self, name, index, undone, avg_V):
        threading.Thread.__init__(self)
        self.name = name
        self.index = index   # 第幾筆資料
        self.undone = undone
        self.avg_V =avg_V
    def run(self):
        beforeTime = time.time()  # 開始時間
        try:
            wb= load_workbook("G-print_3.xlsx")  # 嘗試開啟 excel
        except FileNotFoundError:
            wb = Workbook()    # 建立新的 excel
        ws = wb.create_sheet('Name')  # 建立 sheet

        # excel 排版
        # =========================================================================================================
        ws.append(["", "An0011", "", "", "An0094", "", "", "An0095", "", "", "An0096", "", "", "An0099", "", ""])
        for i in range(2,17,3):
            ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+2)
        ws.append(["Index"] + ["Ranging", "IMU", "Actual"] * 5)
        # =========================================================================================================

        output = []  # 暫存輸出資料
        while Detail_Data.qsize() > 0 or self.undone:  # 當佇列為空且 catchData 已做完則離開
            if Detail_Data.qsize() != 0:
                disData = Detail_Data.get()   # 拿取資料和時間
                arrive_time = Catch_time.get()

                for i in AnchorName:   # 資料排版
                    output.append(self.index)
                    output.append(disData[i]["Ranging"])
                    output.append(disData[i]["IMU"])
                    # 算實際距離 (時間差, 平均速度, CalActual的變數(  , anchor 座標,  ) )
                    output.append(calDis(arrive_time - beforeTime, self.avg_V, carPosition, anchorPositions[i], dire))
                ws.append(output)  # 輸出資料
                self.index += 1  # 當前index
                print(output)
            wb.save('Test.xlsx')  # 存檔